import os
import sqlite3
# import json

import datetime
from datetime import date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


class ReportsManager:
    def __init__(self, folderPath):
        self.primaryDir = folderPath
        self.conn = sqlite3.connect('./marketplace/facebookDB.db')
    def __init__(self):
        self.primaryDir = "C:\\Users\\[USER_PROFILE]\\Desktop\\MPScrubberReports"
        self.conn = sqlite3.connect('./marketplace/facebookDB.db')
    def set_primary_directory(self):
        user_profile = os.environ.get('USERPROFILE')
        # user_profile becomes C:\Users\cj_po
        print(user_profile)
        self.primaryDir = user_profile + "\\Desktop\\MPScrubberReports"
        print(self.primaryDir)

        # Check if the folder exists
        if not os.path.exists(self.primaryDir):
        # Create the folder if it doesn't exist
            os.makedirs(self.primaryDir)
            print("Folder created successfully:", self.primaryDir)
        else:
            print("Folder already exists. No action required.")

    def build_new_report(self, prefBrands, numPostings):
        currDateTime = self.get_current_date_and_time()
        # brandList = self.get_brand_list()
        wb = Workbook()
        ws = wb.active
        
        cursor = self.conn.cursor()
        for brand in prefBrands:
            self.write_brand_data(brand, ws, cursor, numPostings)
        
        # When saving, openpyxl saves a "workbook" not the "worksheet"
        newFile = self.save_new_report(currDateTime, wb)
        self.open_report(newFile)
        self.conn.close()

    def get_current_date_and_time(self):
        today = date.today()
        currDate = today.strftime("%Y-%m-%d")
        time = datetime.datetime.now().time()
        currTime = time.strftime("%H.%M.%S")
        currDateTime = currDate + "-" + currTime
        return currDateTime
    
    def get_brand_list(self):
        cursor = self.conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
        tables = cursor.fetchall()
        brands = [table[0] for table in tables]
        return brands
    
    def write_brand_data(self, brand, worksheet, cursor, numPostings):
         # Fetch data sorted by DatePulled in descending order for the specified brand
        cursor.execute(f"SELECT * FROM {brand} ORDER BY DatePulled DESC")
        # Fetch "numPostings" worth of most recent entries
        data = cursor.fetchmany(numPostings)
        if data:
            # Write brand name in the first cell of a new row (bold and underlined)
            brand_cell_row = worksheet.max_row + 1
            brand_cell = worksheet.cell(row=brand_cell_row, column=1, value=brand)
            brand_cell.font = Font(bold=True, underline='single')

            # Write column titles in the following row (bold)
            column_titles = ["PrimaryKey", "DatePulled", "DatePosted", 
                            "Year", "Price", "Mileage", 
                            "Description", "Location", "Link"]
            title_row = brand_cell_row + 1
            for col_num, title in enumerate(column_titles, start=1):
                title_cell = worksheet.cell(row=title_row, column=col_num, value=title)
                title_cell.font = Font(bold=True)

            data_row_start = title_row + 1
            for row_idx, row in enumerate(data, start=data_row_start):
                for col_idx, value in enumerate(row, start=1):
                    cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                    if column_titles[col_idx - 1] == "Link" and not value.startswith("n/a"):
                        cell.value = "Post Link"
                        cell.hyperlink = value
                        cell.style = "Hyperlink"

        # Skip a line for the next brand's data
        worksheet.append([])

        # Adjust column widths to fit the content
        for col_num in range(1, len(column_titles) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for cell in worksheet[column_letter]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2  # Adding extra space for better readability
            worksheet.column_dimensions[column_letter].width = adjusted_width


    def save_new_report(self, currDateTime, workbook):
        # Excel file type = .xlsx
        reportFileName = "MPReport(" + currDateTime + ").xlsx"
        reportFilePath = self.primaryDir + "\\" + reportFileName
        # If the report doesn't already exist (which it shouldn't)
        if not os.path.exists(reportFilePath):
            print(reportFilePath)
            workbook.save(reportFilePath)
            print("Excel file created successfully:", reportFilePath)
        else:
            pass
        return reportFilePath
    
    def open_report(self, filePath):
        if os.name == 'nt':  # For Windows
            os.startfile(filePath)
        elif os.name == 'posix':  # For macOS and Linux
            s.system(f'open "{file_path}"')
    


   