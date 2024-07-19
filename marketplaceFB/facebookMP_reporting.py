import os
import sqlite3

import datetime
from datetime import date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter



# PrimaryKey TEXT,
# DatePulled TEXT,
# DatePosted TEXT,
# Year INT,
# Price INT,
# Mileage INT,
# Description TEXT,
# Location TEXT,
# Link TEXT


class FB_ExcelReportManager:
    def __init__(self, primaryDir=None):
        self.primaryDir = primaryDir if primaryDir is not None else self.set_primary_directory()
        self.conn = sqlite3.connect('./marketplaceFB/facebookDB.db')

    def set_primary_directory(self):
        user_profile = os.environ.get('USERPROFILE')
        # user_profile becomes C:\Users\cj_po
        primaryDir = user_profile + "\\Desktop\\MPScrubberReports"

        # Check if the folder exists
        if not os.path.exists(primaryDir):
        # Create the folder if it doesn't exist
            os.makedirs(primaryDir)
            print("Folder created successfully:", primaryDir)
        return primaryDir

    def build_new_report(self, prefBrands, numPostings):
        currDateTime = self.get_current_date_and_time()
        # brandList = self.get_brand_list()
        wb = Workbook()
        ws = wb.active
        
        cursor = self.conn.cursor()
        for brand in prefBrands:
            self.write_brand_data(brand, ws, cursor, numPostings)
        
        # When saving, openpyxl saves a "workbook" not the "worksheet"
        newFile = self.save_new_report(currDateTime, wb)
        self.open_report(newFile)
        self.conn.close()

    def get_current_date_and_time(self):
        today = date.today()
        currDate = today.strftime("%Y-%m-%d")
        time = datetime.datetime.now().time()
        currTime = time.strftime("%H.%M.%S")
        currDateTime = currDate + "-" + currTime
        return currDateTime
    
    def get_brand_list(self):
        cursor = self.conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
        tables = cursor.fetchall()
        brands = [table[0] for table in tables]
        return brands
    
    def write_brand_data(self, brand, worksheet, cursor, numPostings):
        sortingType = "DatePulled"
        # Fetch data sorted by DatePulled in descending order for the specified brand
        cursor.execute(f"SELECT * FROM {brand} ORDER BY {sortingType} DESC")
        # Fetch "numPostings" worth of most recent entries
        data = cursor.fetchmany(numPostings)
        if data:
            # Write brand name in the first cell of a new row (bold and underlined)
            brand_cell_row = worksheet.max_row + 1
            brand_cell = worksheet.cell(row=brand_cell_row, column=1, value=brand)
            brand_cell.font = Font(bold=True, underline='single')

            # Write column titles in the following row (bold)
            column_titles = ["PrimaryKey", "DateScraped", "DatePosted", 
                            "Year", "Price", "Mileage", 
                            "Description", "Location", "Link"]
            title_row = brand_cell_row + 1
            for col_num, title in enumerate(column_titles, start=1):
                title_cell = worksheet.cell(row=title_row, column=col_num, value=title)
                title_cell.font = Font(bold=True)

            data_row_start = title_row + 1
            for row_idx, row in enumerate(data, start=data_row_start):
                for col_idx, value in enumerate(row, start=1):
                    cell = worksheet.cell(row=row_idx, column=col_idx, value=value)
                    if column_titles[col_idx - 1] == "Link" and not value.startswith("n/a"):
                        cell.value = "Post Link"
                        cell.hyperlink = value
                        cell.style = "Hyperlink"

        # Skip a line for the next brand's data
        worksheet.append([])

        # Adjust column widths to fit the content
        for col_num in range(1, len(column_titles) + 1):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for cell in worksheet[column_letter]:
                # If the cell has content
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2  # Adding extra space for better readability
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def save_new_report(self, currDateTime, workbook):
        # Excel file type = .xlsx
        reportFileName = "MPReport(" + currDateTime + ").xlsx"
        reportFilePath = self.primaryDir + "\\" + reportFileName
        # If the report doesn't already exist (which it shouldn't)
        if not os.path.exists(reportFilePath):
            workbook.save(reportFilePath)
            print("Report created successfully. Opening now!")
        else:
            pass
        return reportFilePath
    
    def open_report(self, filePath):
        if os.name == 'nt':  # For Windows
            os.startfile(filePath)
        elif os.name == 'posix':  # For macOS and Linux
            s.system(f'open "{file_path}"')
    

import numpy as np
import pandas as pd
import plotly.express as px
from marketplaceFB.facebookMP_database import *

# INTERPRETING TRENDS
# Points Above the Trend Line:
#   Higher Price: A vehicle above the trend line shows that the price is higher than what is expected
#       Potential Implications:
#           Better Condition: The vehicle might be in better condition compared to others with similar mileage.
#           Additional Features: The vehicle may have additional features/options
#           Brand Reputation: The brand/model may have a better reputation/higher demand
#           Market Conditions: Market conditions/seller strategies point to selling at a higher price
# Points Below the Trend Line:
#   Lower Price: A data point below the trend line indicates that the vehicle is priced lower than what is expected based on the trend. This could be due to several factors:
#       Potential Implications:
#           Poor Condition: The vehicle might be in worse condition compared to others with similar mileage.
#           Missing Features: The vehicle may lack features found in similar models
#           Urgent Sale: The seller may be looking for a quick sale/will accept a lower price
#           Market Conditions: Market conditions suck or pricing is competitive for brand/model

class FB_TrendsAnalyzer:
    def __init__(self):
        self.conn = sqlite3.connect('./marketplaceFB/facebookDB.db')
        self.lowestSlope = 0
        self.highestSlope = 0
    def plot_compare_trends_with_data_points(self, db, brands):
        fig = px.scatter(title='Mileage vs. Price Trends Comparison', labels={'Mileage': 'Mileage', 'Price': 'Price'})
        color_cycle = px.colors.qualitative.Plotly
        for i, brand in enumerate(brands):
            data = db.fetch_mileage_and_prices_all(brand)
            df = pd.DataFrame(data, columns=['Mileage', 'Price'])

            # Filter out invalid data, i.e., anything below 0 isn't possible and should be considered 
            #   non-applicable to the data set
            df = df[(df['Mileage'] > 1000) & (df['Price'] > 1000)]

            # Calculate the trend (linear regression)
            slope, intercept = np.polyfit(df['Mileage'], df['Price'], 1)
            trendline = slope * df['Mileage'] + intercept

            # Set color for the brand
            color = color_cycle[i % len(color_cycle)]

            fig.add_scatter(x=df['Mileage'], y=df['Price'], mode='markers', name=f'{brand} Data', marker=dict(color=color))
            fig.add_scatter(x=df['Mileage'], y=trendline, mode='lines', name=f'{brand} Trend', line=dict(color=color))
        # Make graph only show above -1000 on both axis
        # Set max to the highest one plus a little extra
        fig.update_layout(
            xaxis=dict(range=[-1000, df['Mileage'].max() * 1.01]),
            yaxis=dict(range=[-1000, df['Price'].max() * 1.01])
        )
        fig.show()
    def plot_compare_trends_without_data_points(self, db, brands):
        fig = px.scatter(title='Mileage vs. Price (All Brands)', labels={'Mileage': 'Mileage', 'Price': 'Price'})
        color_cycle = px.colors.qualitative.Plotly
        for i, brand in enumerate(brands):
            if db.fetch_total_number_of_entries(brand) < 50:
                continue
            data = db.fetch_mileage_and_prices_all(brand)
            df = pd.DataFrame(data, columns=['Mileage', 'Price'])

            # Filter out invalid data
            df = df[(df['Mileage'] > 1000) & (df['Price'] > 1000)]

            # Calculate the trend (linear regression)
            slope, intercept = np.polyfit(df['Mileage'], df['Price'], 1)
            trendline = slope * df['Mileage'] + intercept

            # Set color for the brand
            color = color_cycle[i % len(color_cycle)]

            # Add only the trendline to the plot
            fig.add_scatter(x=df['Mileage'], y=trendline, mode='lines', name=f'{brand} Trend', line=dict(color=color))

        # Make graph only show above 0 on both axis
        # Set max to the highest one plus a little extra
        fig.update_layout(
            xaxis=dict(range=[-1000, df['Mileage'].max() * 1.01]),
            yaxis=dict(range=[-1000, df['Price'].max() * 1.01])
        )

        fig.show()
    def check_for_good_deal(self, db, brands):
        fig = px.scatter(title='Mileage vs. Price Trends with Good Deals', labels={'Mileage': 'Mileage', 'Price': 'Price'})
        goodDeals = {}
        for brand in brands:
            allData = db.fetch_mileage_and_prices_all(brand)
            newData = db.fetch_mileage_and_prices_most_recent(brand)

            # Create DataFrame and filter out invalid data
            df = pd.DataFrame(allData, columns=['Mileage', 'Price'])
            df = df[(df['Mileage'] > 100) & (df['Price'] > 100)]

            # Calculate the trend (linear regression)
            slope, intercept = np.polyfit(df['Mileage'], df['Price'], 1)
            trendLine = slope * df['Mileage'] + intercept

            # Add only the trendline to the plot
            fig.add_scatter(x=df['Mileage'], y=trendLine, mode='lines', name=f'{brand} Trend')

            for entry in newData:
                primaryKey = entry[0]
                mileage = entry[1]
                price = entry[2]
                if mileage > 0 and price > 0:
                    isGoodDeal = self.is_entry_good_deal(df, mileage, price)
                    if isGoodDeal:
                        fig.add_scatter(x=[mileage], y=[price], mode='markers', name=f'{brand} {primaryKey}', marker=dict(color='red'))
                        goodDeals.update(db.fetch_details_by_primary_key(brand, primaryKey))
        # Show the plot
        fig.update_layout(
            xaxis=dict(range=[0, df['Mileage'].max() * 1.01]),
            yaxis=dict(range=[0, df['Price'].max() * 1.01])
        )
        fig.show()

        return goodDeals

    def is_entry_good_deal(self, df, mileage, price):
        # Calculate the trend (linear regression)
        slope, intercept = np.polyfit(df['Mileage'], df['Price'], 1)
        # Calculate the expected price at the given mileage
        expectedPrice = slope * mileage + intercept
        # Determine if the price is below the trend line but within 10% of the expected price on that trend
        lowerBound = expectedPrice - (expectedPrice * 0.10)
        # Is the posted price within 10%?
        isGoodDeal = lowerBound <= price < expectedPrice
        # Return the result
        return isGoodDeal
   